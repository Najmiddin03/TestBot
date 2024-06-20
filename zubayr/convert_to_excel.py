import asyncio
import aiosqlite
import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.types import InputFile, InlineKeyboardButton, InlineKeyboardMarkup, CallbackQuery
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.utils import executor
from aiogram.dispatcher.filters import Command
from aiogram.utils.callback_data import CallbackData

API_TOKEN = '6587640537:AAGKgU1-Ffb9jRysxGvAfiFgpYIIsBMrYyw'
DATABASE = 'your_database.db'

bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)
dp.middleware.setup(LoggingMiddleware())

callback_data = CallbackData("export", "action")

async def fetch_data_from_db():
    async with aiosqlite.connect(DATABASE) as db:
        async with db.execute("SELECT * FROM user") as cursor:
            columns = [description[0] for description in cursor.description]
            rows = await cursor.fetchall()
    return columns, rows

def create_excel_file(columns, rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the column headers
    for col_num, column_title in enumerate(columns, 1):
        sheet.cell(row=1, column=col_num, value=column_title)

    # Write the data rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    file_path = '/mnt/data/output.xlsx'
    workbook.save(file_path)
    return file_path

@dp.message_handler(Command("start"))
async def send_welcome(message: types.Message):
    keyboard = InlineKeyboardMarkup().add(
        InlineKeyboardButton("Export to Excel", callback_data=callback_data.new(action="convert"))
    )
    await message.reply("Welcome! Click the button below to export data to Excel.", reply_markup=keyboard)

@dp.callback_query_handler(callback_data.filter(action="convert"))
async def export_data(callback_query: CallbackQuery):
    try:
        columns, rows = await fetch_data_from_db()
        file_path = create_excel_file(columns, rows)
        await bot.send_document(callback_query.from_user.id, InputFile(file_path))
        await callback_query.answer("Excel file has been created and sent!")
    except Exception as e:
        await callback_query.answer(f"An error occurred: {e}")

if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
